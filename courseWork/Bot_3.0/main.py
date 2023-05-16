import telebot
import openai
import openpyxl

bot = telebot.TeleBot('')
openai.api_key = ''

WATCHING_LIST_FILE = 'database.xlsx'
WATCHING_LIST_SHEET = 'watching_list'


def load_watching_list(user_id):
    wb = openpyxl.load_workbook(WATCHING_LIST_FILE)
    sheet = wb[WATCHING_LIST_SHEET]
    watching_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == user_id:
            watching_list.append(row[1])
    return watching_list


def add_to_watching_list(user_id, title):
    wb = openpyxl.load_workbook(WATCHING_LIST_FILE)
    sheet = wb[WATCHING_LIST_SHEET]
    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1, value=user_id)
    sheet.cell(row=next_row, column=2, value=title)
    wb.save(WATCHING_LIST_FILE)


def generate_recommendations(watching_list):
    prompt = f"Generate only three anime recommendations based on the following watched titles:\n{', '.join(watching_list)}"
    completions = openai.Completion.create(
        engine="text-davinci-002",
        prompt=prompt,
        max_tokens=100,
        n=3,
        stop=None,
        temperature=0.5,
    )
    recommendations = [choice.text for choice in completions.choices]
    return recommendations


@bot.message_handler(commands=['start'])
def handle_start(message):
    bot.send_message(message.chat.id, 'Welcome to the Anime Recommendation Bot!')
    main_menu(message)


def main_menu(message):
    keyboard = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.row('Watching List', 'Recommend')
    bot.send_message(message.chat.id, 'Please select an option:', reply_markup=keyboard)


@bot.message_handler(func=lambda message: message.text == 'Watching List')
def handle_watching_list(message):
    watching_list = load_watching_list(message.chat.id)
    if len(watching_list) == 0:
        bot.send_message(message.chat.id, "Your watching list is empty.")
    else:
        bot.send_message(message.chat.id, "Your watching list:")
        bot.send_message(message.chat.id, "\n".join(watching_list))
    watching_list_menu(message)


def watching_list_menu(message):
    keyboard = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.row('Add', 'Show')
    keyboard.row('Return Back')
    bot.send_message(message.chat.id, 'Please select an option:', reply_markup=keyboard)
    bot.register_next_step_handler(message, handle_watching_list_menu)

def handle_watching_list_menu(message):
    if message.text == 'Add':
        handle_add_to_watching_list(message)
    elif message.text == 'Show':
        handle_show_watching_list(message)
    elif message.text == 'Return Back':  # new line
        return_to_main_menu(message)  # new line


@bot.message_handler(func=lambda message: message.text == 'Add')
def handle_add_to_watching_list(message):
    bot.send_message(message.chat.id, "Please enter the title of the anime you want to add:")
    bot.register_next_step_handler(message, add_to_watching_list_callback)


def add_to_watching_list_callback(message):
    add_to_watching_list(message.chat.id, message.text)
    bot.send_message(message.chat.id, f"{message.text} has been added to your watching list.")
    watching_list_menu(message)


@bot.message_handler(func=lambda message: message.text == 'Show')
def handle_show_watching_list(message):
    watching_list = load_watching_list(message.chat.id)
    if len(watching_list) == 0:
        bot.send_message(message.chat.id, "Your watching list is empty.")
    else:
        bot.send_message(message.chat.id, "Your watching list:")
        bot.send_message(message.chat.id, "\n".join(watching_list))
    watching_list_menu(message)


@bot.message_handler(func=lambda message: message.text == 'Recommend')
def handle_recommend(message):
    watching_list = load_watching_list(message.chat.id)
    if len(watching_list) == 0:
        bot.send_message(message.chat.id, "Your watching list is empty. Please add some titles first.")
        watching_list_menu(message)
    else:
        recommendations = generate_recommendations(watching_list)
        bot.send_message(message.chat.id, "Here are some anime recommendations:")
        bot.send_message(message.chat.id, "\n".join(recommendations))
        refresh_menu(message)


def refresh_menu(message):
    keyboard = telebot.types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.row('Refresh')
    keyboard.row('Return Back')
    bot.send_message(message.chat.id, 'Please select an option:', reply_markup=keyboard)
    bot.register_next_step_handler(message, handle_refresh_list_menu)


@bot.message_handler(func=lambda message: message.text == 'Refresh')
def handle_refresh(message):
    watching_list = load_watching_list(message.chat.id)
    recommendations = generate_recommendations(watching_list)
    bot.send_message(message.chat.id, "Here are some updated anime recommendations:")
    bot.send_message(message.chat.id, "\n".join(recommendations))
    refresh_menu(message)

def handle_refresh_list_menu(message):
    if message.text == 'Refresh':
        handle_refresh(message)
    elif message.text == 'Return Back':
        return_to_main_menu(message)

def return_to_main_menu(message):
    bot.send_message(message.chat.id, "Returning to main menu...")
    main_menu(message)


bot.polling()
